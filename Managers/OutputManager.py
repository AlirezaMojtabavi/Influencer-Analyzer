from datetime import datetime, timedelta, timezone, date
import xlsxwriter
import openpyxl
import pytz
from copy import copy
from os import path, makedirs
from openpyxl.worksheet.table import Table

from Repositories.InfluencerRepository import InfluencerRepository as InflRepo, Privacy
from Repositories.AdvertiseRepository import AdvertiseRepository as AdvRepo, AdvertiseStatus


class OutputManager:
    def __init__(self):
        pass

    def create_xlsx_result(self, influencer_username):
        infl_id = InflRepo.check_influencer_exists(influencer_username)
        if not infl_id:
            return

        address = "OutputFiles/" + influencer_username + ".xlsx"
        workbook = xlsxwriter.Workbook(address)
        str_format = workbook.add_format({'font_name': 'B Homa', 'font_size': 14})
        str_format.set_center_across()
        str_format.set_align('center')
        int_format = workbook.add_format({'num_format': "Number"})
        int_format.set_center_across()

        influencer_sheet = workbook.add_worksheet("influencer_datas")
        influencer_sheet.right_to_left()
        influencer_sheet.write(0, 0, "شناسه اینفلوئنسر", str_format)
        influencer_sheet.write(0, 1, InflRepo.get_username(infl_id), str_format)
        influencer_sheet.write(1, 0, "نام اکانت", str_format)
        influencer_sheet.write(1, 1, InflRepo.get_full_name(infl_id), str_format)
        influencer_sheet.write(2, 0, "بیوگرافی", str_format)
        influencer_sheet.write(2, 1, InflRepo.get_bio(infl_id), str_format)
        influencer_sheet.write(3, 0, "تعداد فالور", str_format)
        influencer_sheet.write(3, 1, InflRepo.get_follower(infl_id), int_format)
        influencer_sheet.write(4, 0, "تعداد پست", str_format)
        influencer_sheet.write(4, 1, InflRepo.get_media_count(infl_id), int_format)
        influencer_sheet.write(5, 0, "میانگین لایک", str_format)
        influencer_sheet.write(5, 1, InflRepo.get_AVG_Like(infl_id), int_format)
        influencer_sheet.write(6, 0, "میانگین کامنت", str_format)
        influencer_sheet.write(6, 1, InflRepo.get_AVG_Comment(infl_id), int_format)
        influencer_sheet.write(7, 0, "EngagementRate", str_format)
        influencer_sheet.write(7, 1, InflRepo.get_engagement_rate(infl_id), int_format)

        mention_account_sheet = workbook.add_worksheet("mention_accounts")
        mention_account_sheet.right_to_left()
        date_format = workbook.add_format(
            {'num_format': '[$-fa-IR,16]yyyy/mm/dd', 'font_name': 'B Homa', 'font_size': 8})
        time_format = workbook.add_format({'num_format': 'h:mm'})

        mention_account_sheet.write(0, 0, "ردیف", str_format)
        mention_account_sheet.write(0, 1, "تاریخ تبلیغ", str_format)
        mention_account_sheet.write(0, 2, "ساعت تبلیغ", str_format)
        mention_account_sheet.write(0, 3, "شناسه پیج معرفی شده", str_format)
        mention_account_sheet.write(0, 4, "نام پیج معرفی شده", str_format)
        mention_account_sheet.write(0, 5, "بیوگرافی پیج معرفی شده", str_format)
        mention_account_sheet.write(0, 6, "تعداد پست", str_format)
        mention_account_sheet.write(0, 7, "تعداد استوری تبلیغ", str_format)
        mention_account_sheet.write(0, 8, "تعداد فالور قبل تبلیغ", str_format)
        mention_account_sheet.write(0, 9, "تعداد فالور بعد از 1 ساعت تبلیغ", str_format)
        mention_account_sheet.write(0, 10, "تعداد فالور بعد از 2 ساعت تبلیغ", str_format)
        mention_account_sheet.write(0, 11, "تعداد فالور بعد از 12 ساعت تبلیغ", str_format)
        mention_account_sheet.write(0, 12, "تعداد فالور بعد از 24 ساعت تبلیغ", str_format)
        mention_account_sheet.write(0, 13, "بازخورد", str_format)
        mention_account_sheet.write(0, 14, "وضعیت اکانت", str_format)
        mention_account_sheet.write(0, 15, "وضعیت تبلیغ", str_format)
        workbook.close()

        #### copy style
        workbook1 = openpyxl.load_workbook(address)
        inf_account_sheet = workbook1["influencer_datas"]
        men_account_sheet = workbook1["mention_accounts"]

        workbook_t = openpyxl.load_workbook("InputFiles/template.xlsx")
        inf_account_sheet_t = workbook_t["influencer_datas"]
        men_account_sheet_t = workbook_t["mention_accounts"]
        inf_account_sheet.column_dimensions["B"].width = 60.0
        inf_account_sheet.column_dimensions["A"].width = 40.0
        for i in range(1, 9):
            newcell = inf_account_sheet.cell(i, 1)
            cell = inf_account_sheet_t.cell(i, 1)
            newcell.value = cell.value
            self.format_copy(newcell, cell)
        for i in range(1, 9):
            newcell = inf_account_sheet.cell(i, 2)
            cell = inf_account_sheet_t.cell(i, 2)
            self.format_copy(newcell, cell)
        # mention account sheet
        men_account_sheet.column_dimensions["A"].width = 20.0
        men_account_sheet.column_dimensions["B"].width = 20.0
        men_account_sheet.column_dimensions["C"].width = 20.0
        men_account_sheet.column_dimensions["D"].width = 30.0
        men_account_sheet.column_dimensions["E"].width = 25.0
        men_account_sheet.column_dimensions["F"].width = 80.0
        men_account_sheet.column_dimensions["G"].width = 15.0
        men_account_sheet.column_dimensions["H"].width = 15.0
        men_account_sheet.column_dimensions["I"].width = 20.0
        men_account_sheet.column_dimensions["J"].width = 20.0
        men_account_sheet.column_dimensions["K"].width = 20.0
        men_account_sheet.column_dimensions["L"].width = 20.0
        men_account_sheet.column_dimensions["M"].width = 20.0
        men_account_sheet.column_dimensions["N"].width = 20.0
        men_account_sheet.column_dimensions["O"].width = 20.0
        men_account_sheet.column_dimensions["P"].width = 20.0
        men_account_sheet.column_dimensions["Q"].width = 20.0

        for j in range(1, 18):
            newcell1 = men_account_sheet.cell(1, j)
            cell1 = men_account_sheet_t.cell(1, j)
            newcell1.value = cell1.value
            self.format_copy(newcell1, cell1)

        workbook1.save(address)

    def format_copy(self, nc, c):
        nc.font = copy(c.font)
        nc.border = copy(c.border)
        nc.fill = copy(c.fill)
        nc.number_format = copy(c.number_format)
        nc.protection = copy(c.protection)
        nc.alignment = copy(c.alignment)

    def create_xlsx_report(self):
        for infl_id in InflRepo.get_all_infl_ids():
            file_name = InflRepo.get_username(infl_id) + ".xlsx"
            if not path.isdir("OutputFiles"):
                makedirs("OutputFiles")

            self.create_xlsx_result(InflRepo.get_username(infl_id))

            address = "OutputFiles/" + file_name
            workbook = openpyxl.load_workbook(address)
            mention_account_sheet = workbook["mention_accounts"]

            new_row = mention_account_sheet.max_row + 1
            columns_number = mention_account_sheet.max_column
            workbook_t = openpyxl.load_workbook("InputFiles/template.xlsx")

            template_file = openpyxl.load_workbook("InputFiles/template.xlsx")
            template_sheet = template_file["mention_accounts"]
            template_line = template_sheet.max_row

            if len(mention_account_sheet.tables) != 0:
                del mention_account_sheet.tables["Table1"]
            ##############################
            adv_ids = AdvRepo.get_all_infl_advs(infl_id)
            for adv_id in adv_ids:
                if (new_row % 2) == 0:
                    template_line = 2
                else:
                    template_line = 3

                for i in range(1, columns_number + 1):
                    nc = mention_account_sheet.cell(new_row, i)
                    cc = template_sheet.cell(template_line, i)
                    self.format_copy(nc, cc)
                mention_account_name = AdvRepo.get_mentioned_username(adv_id)
                story_date = AdvRepo.get_start_datetime(adv_id).astimezone(pytz.timezone('Asia/Tehran')).date()
                changed_date = datetime.strptime(story_date.strftime("%Y-%m-%d"), "%Y-%m-%d")
                mention_account_sheet.cell(new_row, 2).value = changed_date
                mention_account_sheet.cell(new_row, 2).number_format = "[$-fa-IR,16]yyyy/mm/dd"
                mention_account_sheet.cell(new_row, 1).value = new_row - 1

                story_time = AdvRepo.get_start_datetime(adv_id).astimezone(pytz.timezone('Asia/Tehran')).time()
                changed_time = datetime.strptime(story_time.strftime("%H-%M"), "%H-%M")
                mention_account_sheet.cell(new_row, 3).value = changed_time
                mention_account_sheet.cell(new_row, 3).number_format = 'h:mm'

                mention_account_sheet.cell(new_row, 4).value = mention_account_name
                mention_account_sheet.cell(new_row, 5).value = AdvRepo.get_mentioned_username(adv_id)
                mention_account_sheet.cell(new_row, 6).value = AdvRepo.get_mentioned_bio(adv_id)
                mention_account_sheet.cell(new_row, 7).value = AdvRepo.get_mentioned_media_count(adv_id)
                mention_account_sheet.cell(new_row, 8).value = AdvRepo.get_mentioned_story_count(adv_id)
                mention_account_sheet.cell(new_row, 9).value = AdvRepo.get_follower_before(adv_id)
                mention_account_sheet.cell(new_row, 10).value = AdvRepo.get_follower_after_1h(adv_id)
                mention_account_sheet.cell(new_row, 11).value = AdvRepo.get_follower_after_2h(adv_id)
                mention_account_sheet.cell(new_row, 12).value = AdvRepo.get_follower_after_12h(adv_id)
                mention_account_sheet.cell(new_row, 13).value = AdvRepo.get_follower_after_24h(adv_id)
                mention_account_sheet.cell(new_row,
                                           14).value = AdvRepo.get_follower_after_24h(adv_id) - \
                                                       AdvRepo.get_follower_before(adv_id)

                if AdvRepo.get_mentioned_privacy(adv_id) == Privacy.Public:
                    mention_account_sheet.cell(new_row, 15).value = "عمومی"  # public
                else:
                    mention_account_sheet.cell(new_row, 15).value = "خصوصی"

                if AdvRepo.get_mentioned_status(adv_id) == AdvertiseStatus.Open:
                    mention_account_sheet.cell(new_row, 16).value = "باز"
                else:
                    mention_account_sheet.cell(new_row, 16).value = "بسته"
                new_row = new_row + 1

            endline = mention_account_sheet.max_row
            refs = "A1:P" + str(endline)

            tab = Table(displayName="Table1", ref=refs)
            mention_account_sheet.add_table(tab)
            workbook.save(address)
