from instagrapi import Client
import openpyxl
from time import sleep
from instaloader import instaloader
import random
from datetime import datetime, timezone
from Repositories.EventRepository import EventRepository
from Repositories.BotRepository import BotRepository, BotAccountStatus, LastLoginStatus
from Models.Event import EventName, EventCategory, RobotAPI, EventType
from instagrapi.types import (
    Account, Collection, Comment, DirectMessage, DirectThread, Hashtag, Highlight, Location, Media, MediaOembed,
    Share, Story, StoryLink, StoryMedia, StoryMention, StorySticker, User, UserShort, Usertag,
)

from instagrapi.exceptions import (
    BadPassword,
    ChallengeRequired,
    FeedbackRequired,
    LoginRequired,
    PleaseWaitFewMinutes,
    RecaptchaChallengeForm,
    ReloginAttemptExceeded,
    SelectContactPointRecoveryForm,
)


class Robot:
    def __init__(self, accounts_path):
        self.accounts_path = accounts_path
        self.grapi_username = ""
        self.grapi_password = ""
        self.influencer_array = []
        self.bots_account_array = []
        self.grapi_account_id = None
        self.start_cycle_time_engine = None
        self.set_authentication(accounts_path)
        self.grapi_engine = None
        self.engine_flag = None
        self.loader_account_id = None
        self.engine_flag = self.set_grapi_engine()
        self.loader_username = ""
        self.loader_password = ""
        self.loader_engine = None
        if self.engine_flag:
            self.prepare_loader_engine()
            self.start_cycle_time_engine = datetime.now(timezone.utc)

    def set_authentication(self, accounts_path):
        self.update_bots_account_list(accounts_path)
        self.grapi_username = self.bots_account_array[0][1]
        self.grapi_password = self.bots_account_array[0][2]
        self.grapi_account_id = 1

    def set_grapi_engine(self):
        self.grapi_engine = Client()
        try:
            self.grapi_engine.login(self.grapi_username, self.grapi_password)
            EventRepository.register_event(EventType.Log, EventCategory.Authentication,
                                           EventName.Login_Succeeded, "Robot::set_grapi_engine",
                                           robot_api=RobotAPI.Instagrapi, bot_id=self.get_grapi_acc_id(),
                                           content="Robot logged")
            BotRepository.update_bot_content(self.grapi_username, BotAccountStatus.InstaGrapi,
                                             last_login_status=LastLoginStatus.Successful, message=' ')
            return True
        except Exception as e:
            EventRepository.register_event(EventType.Warning, EventCategory.Authentication,
                                           EventName.Login_Failed, "Robot::set_grapi_engine",
                                           robot_api=RobotAPI.Instagrapi, bot_id=self.get_grapi_acc_id(),
                                           content=str(e))
            BotRepository.update_bot_content(self.grapi_username, BotAccountStatus.InstaGrapi,
                                             last_login_status=LastLoginStatus.Failed, message=str(e))
            login_flag = self.change_account(message=str(e))
            if login_flag:
                return True
            else:
                EventRepository.register_event(EventType.Error, EventCategory.Authentication,
                                               EventName.All_Changing_Account_Failed, "Robot::set_grapi_engine",
                                               robot_api=RobotAPI.Instagrapi, bot_id=self.get_grapi_acc_id())
                return False

    def prepare_loader_engine(self):
        if self.grapi_account_id >= len(self.bots_account_array):
            self.loader_account_id = 1
        else:
            self.loader_account_id = self.grapi_account_id + 1
        self.loader_username = self.bots_account_array[self.loader_account_id - 1][1]
        self.loader_password = self.bots_account_array[self.loader_account_id - 1][2]
        self.loader_engine = instaloader.Instaloader()
        try:
            self.loader_engine.login(self.loader_username, self.loader_password)
            EventRepository.register_event(EventType.Log, EventCategory.Authentication,
                                           EventName.Login_Succeeded, "Robot::prepare_loader_engine",
                                           robot_api=RobotAPI.Instaloader, bot_id=self.get_loader_acc_id(),
                                           content="Robot logged")
            BotRepository.update_bot_content(self.loader_username, BotAccountStatus.Instaloader,
                                             last_login_status=LastLoginStatus.Successful, message=' ')
            return True
        except Exception as e:
            EventRepository.register_event(EventType.Warning, EventCategory.Authentication,
                                           EventName.Login_Failed, "Robot::prepare_loader_engine",
                                           robot_api=RobotAPI.Instaloader, bot_id=self.get_loader_acc_id(),
                                           content=str(e))
            BotRepository.update_bot_content(self.loader_username, BotAccountStatus.Instaloader,
                                             last_login_status=LastLoginStatus.Failed, message=str(e))
            login_flag = self.change_account("loader", message=str(e))
            if login_flag:
                return True
            else:
                EventRepository.register_event(EventType.Error, EventCategory.Authentication,
                                               EventName.All_Changing_Account_Failed, "Robot::prepare_loader_engine",
                                               robot_api=RobotAPI.Instaloader, bot_id=self.get_loader_acc_id())
                return False

    def get_engine(self):
        return self.grapi_engine

    def update_influencer_list(self, influencer_list_path):
        workbook = openpyxl.load_workbook(influencer_list_path)
        sheet = workbook.active
        for row in range(2, 9000):
            if sheet.cell(row, 1).value:
                row_id = int(sheet.cell(row, 1).value)
                user = sheet.cell(row, 2).value
                start_date = sheet.cell(row, 3).value
                end_date = sheet.cell(row, 4).value
                price = sheet.cell(row, 5).value
                if price is not None:
                    self.influencer_array.append([row_id, user, start_date.astimezone(timezone.utc),
                                                  end_date.astimezone(timezone.utc), price])
                else:
                    self.influencer_array.append([row_id, user, start_date.astimezone(timezone.utc),
                                                  end_date.astimezone(timezone.utc)])
            else:
                break

    def update_bots_account_list(self, accounts_path=None):
        self.bots_account_array.clear()
        accounts_path = self.accounts_path
        workbook = openpyxl.load_workbook(accounts_path)
        sheet = workbook.active
        for row in range(2, sheet.max_row):
            if sheet.cell(row, 1).value:
                row_id = int(sheet.cell(row, 1).value)
                username = str(sheet.cell(row, 2).value)
                password = str(sheet.cell(row, 3).value)
                phone_num = str(sheet.cell(row, 4).value)
                self.bots_account_array.append([row_id, username, password, phone_num])
                BotRepository.update_bot_authentication(username, password, phone_num)
            else:
                break
        BotRepository.update_inactive_bots(self.bots_account_array)

    def get_influencer_array(self):
        return self.influencer_array

    def change_grapi_account(self, message=None):
        if message and message is not None:
            BotRepository.update_bot_content(self.grapi_username, BotAccountStatus.Standby, message=message)
        else:
            BotRepository.update_bot_content(self.grapi_username, BotAccountStatus.Standby)

        if self.grapi_account_id >= len(self.bots_account_array) and self.loader_account_id != 1:
            self.grapi_account_id = 1
        elif self.grapi_account_id >= len(self.bots_account_array) and self.loader_account_id == 1:
            self.grapi_account_id = self.loader_account_id + 1
        elif self.grapi_account_id + 1 == self.loader_account_id and \
                self.loader_account_id >= len(self.bots_account_array):
            self.grapi_account_id = 1
        elif self.grapi_account_id + 1 == self.loader_account_id and \
                self.loader_account_id < len(self.bots_account_array):
            self.grapi_account_id += 2
        else:
            self.grapi_account_id += 1

        self.grapi_username = self.bots_account_array[self.grapi_account_id - 1][1]
        self.grapi_password = self.bots_account_array[self.grapi_account_id - 1][2]
        try:
            sleep(random.randint(2, 4))
            self.grapi_engine.logout()
            if message and message is not None:
                BotRepository.update_bot_content(self.grapi_username, BotAccountStatus.Standby, message=message)
            else:
                BotRepository.update_bot_content(self.grapi_username, BotAccountStatus.Standby)
            sleep(random.randint(9, 16))
            self.grapi_engine = Client()
            self.grapi_engine.username = self.grapi_username
            self.grapi_engine.password = self.grapi_password
            self.grapi_engine.login(self.grapi_username, self.grapi_password)
            EventRepository.register_event(EventType.Log, EventCategory.Authentication,
                                           EventName.Login_Succeeded, "Robot::change_grapi_account",
                                           robot_api=RobotAPI.Instagrapi, bot_id=self.get_grapi_acc_id(),
                                           content="Robot logged")
            BotRepository.update_bot_content(self.grapi_username, BotAccountStatus.InstaGrapi,
                                             LastLoginStatus.Successful, message=' ')
            self.start_cycle_time_engine = datetime.now(timezone.utc)
            sleep(random.randint(3, 5))
            return True
        except Exception as e:
            EventRepository.register_event(EventType.Warning, EventCategory.Authentication,
                                           EventName.Login_Failed, "Robot::change_grapi_account",
                                           robot_api=RobotAPI.Instagrapi, bot_id=self.get_grapi_acc_id(),
                                           content=str(e))
            BotRepository.update_bot_content(self.grapi_username, BotAccountStatus.InstaGrapi, LastLoginStatus.Failed,
                                             str(e))
            return False

    def change_account(self, engine_type="grapi", message=None):
        self.update_bots_account_list()
        login_trying_number = 0
        login_flag = True
        if engine_type == "grapi":
            while login_trying_number < self.get_reserved_accounts_count():
                login_trying_number += 1
                if self.change_grapi_account(message):
                    return True
                else:
                    login_flag = False
                    # sleep(25)
                if login_trying_number > 4:
                    EventRepository.register_event(EventType.Warning, EventCategory.Authentication,
                                                   EventName.Login_Failed, "Robot::change_account",
                                                   robot_api=RobotAPI.Instagrapi, bot_id=self.get_grapi_acc_id(),
                                                   content="Changing account trying number: " + str(
                                                       login_trying_number))

            if not login_flag:
                return False
        else:
            while login_trying_number < self.get_reserved_accounts_count():
                login_trying_number += 1
                if self.change_loader_account(message):
                    return True
                else:
                    login_flag = False
                    # sleep(25)
                if login_trying_number > 4:
                    EventRepository.register_event(EventType.Warning, EventCategory.Authentication,
                                                   EventName.Login_Failed, "Robot::change_account",
                                                   robot_api=RobotAPI.Instaloader, bot_id=self.get_loader_acc_id(),
                                                   content="Changing account trying number: " + str(
                                                       login_trying_number))
            if not login_flag:
                return False

    def change_loader_account(self, message=None):
        if message and message is not None:
            BotRepository.update_bot_content(self.loader_username, BotAccountStatus.Standby, message=message)
        else:
            BotRepository.update_bot_content(self.loader_username, BotAccountStatus.Standby)
        if self.loader_account_id >= len(self.bots_account_array) and self.grapi_account_id != 1:
            self.loader_account_id = 1
        elif self.loader_account_id >= len(self.bots_account_array) and self.grapi_account_id == 1:
            self.loader_account_id = self.grapi_account_id + 1
        elif self.loader_account_id + 1 == self.grapi_account_id and \
                self.grapi_account_id >= len(self.bots_account_array):
            self.loader_account_id = 1
        elif self.loader_account_id + 1 == self.grapi_account_id and \
                self.grapi_account_id < len(self.bots_account_array):
            self.loader_account_id += 2
        else:
            self.loader_account_id += 1

        self.loader_username = self.bots_account_array[self.loader_account_id - 1][1]
        self.loader_password = self.bots_account_array[self.loader_account_id - 1][2]
        try:
            sleep(random.randint(2, 7))
            if message and message is not None:
                BotRepository.update_bot_content(self.loader_username, BotAccountStatus.Standby, message=message)
            else:
                BotRepository.update_bot_content(self.loader_username, BotAccountStatus.Standby)
            self.loader_engine = instaloader.Instaloader()
            self.loader_engine.login(self.loader_username, self.loader_password)
            EventRepository.register_event(EventType.Log, EventCategory.Authentication,
                                           EventName.Login_Succeeded, "Robot::change_loader_account",
                                           robot_api=RobotAPI.Instaloader, bot_id=self.get_loader_acc_id(),
                                           content="Robot logged")
            BotRepository.update_bot_content(self.loader_username, BotAccountStatus.Instaloader,
                                             LastLoginStatus.Successful, message=' ')

            sleep(random.randint(4, 17))
            return True
        except Exception as e:
            EventRepository.register_event(EventType.Warning, EventCategory.Authentication,
                                           EventName.Login_Failed, "Robot::change_loader_account",
                                           robot_api=RobotAPI.Instaloader, bot_id=self.get_loader_acc_id(),
                                           content=str(e))
            BotRepository.update_bot_content(self.loader_username, BotAccountStatus.Instaloader, LastLoginStatus.Failed,
                                             str(e))
            return False

    def get_reserved_accounts_count(self):
        return len(self.bots_account_array)

    def get_login_flag(self):
        return self.engine_flag

    def get_follower_count(self, mentioned_username, adv_id):
        try:
            mentioned_profile = self.loader_engine.check_profile_id(mentioned_username)
            followers_count = mentioned_profile.followers
            return followers_count
        except Exception as e:
            EventRepository.register_event(EventType.Error, EventCategory.Reading_Media,
                                           EventName.Reading_Profile_Failed, "Robot::get_follower_count",
                                           robot_api=RobotAPI.Instaloader, target_account=mentioned_username,
                                           bot_id=self.get_loader_acc_id(), content=str(e))
            login_flag = self.change_account("instaloader")
            if login_flag:
                try:
                    mentioned_profile = self.loader_engine.check_profile_id(mentioned_username)
                    followers_count = mentioned_profile.followers
                    return followers_count
                except Exception as e:
                    EventRepository.register_event(EventType.Error, EventCategory.Reading_Media,
                                                   EventName.Reading_Profile_Failed,
                                                   "Robot::get_follower_count",
                                                   robot_api=RobotAPI.Instaloader, target_account=mentioned_username,
                                                   bot_id=self.get_loader_acc_id(), content=str(e))
                    try:
                        mentioned_id = self.grapi_engine.user_id_from_username(mentioned_username)
                        mentioned_profile = self.grapi_engine.user_info(mentioned_id)
                        return mentioned_profile.follower_count
                    except Exception as e:
                        EventRepository.register_event(EventType.Error, EventCategory.Reading_Media,
                                                       EventName.Reading_Profile_Failed,
                                                       "Robot::get_follower_count",
                                                       robot_api=RobotAPI.Instagrapi, target_account=mentioned_username,
                                                       bot_id=self.get_grapi_acc_id(), content=str(e))
                        return None

            else:
                EventRepository.register_event(EventType.Error, EventCategory.Authentication,
                                               EventName.All_Changing_Account_Failed, "Robot::get_follower_count",
                                               robot_api=RobotAPI.Instaloader, bot_id=self.get_loader_acc_id())
                return False

    def get_start_cycle_time_engine(self):
        return self.start_cycle_time_engine

    def update_cycle_time_engine(self):
        self.start_cycle_time_engine = datetime.now(timezone.utc)

    def get_grapi_acc_id(self):
        return BotRepository.get_bot_id(self.grapi_username)

    def get_loader_acc_id(self):
        return BotRepository.get_bot_id(self.loader_username)
