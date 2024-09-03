from Models.Influencer import Monitoring_Status
from Models.Account import Account, Privacy
from Models.Advertise import Advertise_Status
from instagrapi import Client
from datetime import datetime, timedelta, timezone, date
import xlsxwriter
import openpyxl
import pytz
from copy import copy
from os import path, makedirs
from openpyxl.worksheet.table import Table
import random
from time import sleep

from instagrapi.exceptions import PleaseWaitFewMinutes, LoginRequired
from Repositories.Influencer_repository import InfluencerRepository
from Repositories.Advertise_Repository import AdvertiseRepository
from Repositories.EventRepository import EventRepository
from Models.Event import EventName, EventCategory, RobotAPI, EventType


class InfluencerContainer:
    def __init__(self, bot):
        self.IA_engine = bot
        self.infl_repo = InfluencerRepository()
        self.adv_repo = AdvertiseRepository()

    def get_active_influencers_id(self):
        return self.infl_repo.get_active_influencers()

    def update_inactive_influencers(self):
        self.infl_repo.update_inactive_influencers()

    def update_story_table(self):
        self.adv_repo.update_story_table()

    def check_infl_array_list(self, inf_array_list):
        for influencer_item in inf_array_list:
            influencer_username = influencer_item[1]
            start_monitoring_Datetime = influencer_item[2]
            finish_monitoring_Datetime = influencer_item[3]
            if len(influencer_item) > 4:
                price = influencer_item[4]
            else:
                price = None

            influencer_id = self.infl_repo.check_influencer_exists(influencer_username)
            if influencer_id is None:
                if price is not None:
                    self.create_influencer(influencer_username, start_monitoring_Datetime,
                                           finish_monitoring_Datetime, price)
                else:
                    self.create_influencer(influencer_username, start_monitoring_Datetime,
                                           finish_monitoring_Datetime)
            else:
                self.infl_repo.set_monitoring_datetime(influencer_id, start_monitoring_Datetime,
                                                       finish_monitoring_Datetime)
                if price is not None:
                    self.infl_repo.update_price(influencer_id, price)

    def read_stories(self, infl_id, reading_stories_flag):
        infl_username = self.infl_repo.get_username(infl_id)
        infl_user_id = self.infl_repo.get_user_id(infl_id)
        sleep(random.randint(22, 40))
        try:
            stories = self.IA_engine.grapi_engine.user_stories(infl_user_id)
            EventRepository.register_event(EventType.Log, EventCategory.Reading_Media, EventName.Reading_Story,
                                           "InfluencerContainer::read_stories", robot_api=RobotAPI.Instagrapi,
                                           target_account=infl_username, bot_id=self.IA_engine.get_grapi_acc_id())
        except PleaseWaitFewMinutes as e:
            EventRepository.register_event(EventType.Warning, EventCategory.Reading_Media,
                                           EventName.Reading_Story_Failed,
                                           "InfluencerContainer::read_stories-PleaseWaitFewMinutes",
                                           robot_api=RobotAPI.Instagrapi, target_account=infl_username,
                                           bot_id=self.IA_engine.get_grapi_acc_id(), content=str(e))
            sleep(random.randint(301, 367))  # Wait for 5-6
            self.IA_engine.grapi_engine.logger.exception(e)
            relogin_flag = self.IA_engine.relogin_grapi("InfluencerContainer::read_stories-PleaseWaitFewMinutes",
                                                        infl_username)
            if relogin_flag:
                try:
                    stories = self.IA_engine.grapi_engine.user_stories(infl_user_id)
                    EventRepository.register_event(EventType.Log, EventCategory.Reading_Media, EventName.Reading_Story,
                                                   "InfluencerContainer::read_stories", robot_api=RobotAPI.Instagrapi,
                                                   target_account=infl_username,
                                                   bot_id=self.IA_engine.get_grapi_acc_id())
                except Exception as e:
                    EventRepository.register_event(EventType.Warning, EventCategory.Reading_Media,
                                                   EventName.Reading_Story_Failed,
                                                   "InfluencerContainer::read_stories-PleaseWaitFewMinutes",
                                                   robot_api=RobotAPI.Instagrapi, target_account=infl_username,
                                                   bot_id=self.IA_engine.get_grapi_acc_id(), content=str(e))
                    return None, False
            else:
                login_flag = self.IA_engine.change_account()
                if login_flag:
                    try:
                        stories = self.IA_engine.grapi_engine.user_stories(infl_user_id)
                        EventRepository.register_event(EventType.Log, EventCategory.Reading_Media,
                                                       EventName.Reading_Story, "InfluencerContainer::read_stories"
                                                                                "-PleaseWaitFewMinutes",
                                                       robot_api=RobotAPI.Instagrapi, target_account=infl_username,
                                                       bot_id=self.IA_engine.get_grapi_acc_id())
                    except Exception as e:
                        EventRepository.register_event(EventType.Warning, EventCategory.Reading_Media,
                                                       EventName.Reading_Story_Failed,
                                                       "InfluencerContainer::read_stories-PleaseWaitFewMinutes",
                                                       robot_api=RobotAPI.Instagrapi, target_account=infl_username,
                                                       bot_id=self.IA_engine.get_grapi_acc_id(), content=str(e))
                        return None, False
                else:
                    EventRepository.register_event(EventType.Error, EventCategory.Authentication,
                                                   EventName.All_Changing_Account_Failed,
                                                   "InfluencerContainer::read_stories-PleaseWaitFewMinutes",
                                                   robot_api=RobotAPI.Instagrapi, target_account=infl_username,
                                                   bot_id=self.IA_engine.get_grapi_acc_id())
                    return False, False

        except LoginRequired as e:
            EventRepository.register_event(EventType.Warning, EventCategory.Reading_Media,
                                           EventName.Reading_Story_Failed,
                                           "InfluencerContainer::read_stories-LoginRequired",
                                           robot_api=RobotAPI.Instagrapi, target_account=infl_username,
                                           bot_id=self.IA_engine.get_grapi_acc_id(), content=str(e))

            self.IA_engine.grapi_engine.logger.exception(e)
            sleep(random.randint(30, 60))
            relogin_flag = self.IA_engine.relogin_grapi.relogin("InfluencerContainer::read_stories-LoginRequired",
                                                                infl_username)
            # EventRepository.register_event(EventType.Log, EventCategory.Authentication,
            #                                EventName.Login_Succeeded,
            #                                "InfluencerContainer::read_stories-LoginRequired",
            #                                robot_api=RobotAPI.Instagrapi, target_account=infl_username,
            #                                bot_id=self.IA_engine.get_grapi_acc_id(),
            #                                content="relogged in successfully")
            if relogin_flag:
                try:
                    stories = self.IA_engine.grapi_engine.user_stories(infl_user_id)
                    EventRepository.register_event(EventType.Log, EventCategory.Reading_Media,
                                                   EventName.Reading_Story, "InfluencerContainer::read_stories-"
                                                                            "LoginRequired",
                                                   robot_api=RobotAPI.Instagrapi, target_account=infl_username,
                                                   bot_id=self.IA_engine.get_grapi_acc_id())
                except Exception as e:
                    EventRepository.register_event(EventType.Warning, EventCategory.Reading_Media,
                                                   EventName.Reading_Story_Failed,
                                                   "InfluencerContainer::read_stories-LoginRequired",
                                                   robot_api=RobotAPI.Instagrapi, target_account=infl_username,
                                                   bot_id=self.IA_engine.get_grapi_acc_id(), content=str(e))
                    return None, False
            else:
                login_flag = self.IA_engine.change_account()
                if login_flag:
                    try:
                        stories = self.IA_engine.grapi_engine.user_stories(infl_user_id)
                        EventRepository.register_event(EventType.Log, EventCategory.Reading_Media,
                                                       EventName.Reading_Story, "InfluencerContainer::read_stories-"
                                                                                "LoginRequired",
                                                       robot_api=RobotAPI.Instagrapi, target_account=infl_username,
                                                       bot_id=self.IA_engine.get_grapi_acc_id())
                    except Exception as e:
                        EventRepository.register_event(EventType.Warning, EventCategory.Reading_Media,
                                                       EventName.Reading_Story_Failed,
                                                       "InfluencerContainer::read_stories-LoginRequired",
                                                       robot_api=RobotAPI.Instagrapi, target_account=infl_username,
                                                       bot_id=self.IA_engine.get_grapi_acc_id(), content=str(e))
                        return None, False
                else:
                    EventRepository.register_event(EventType.Error, EventCategory.Authentication,
                                                   EventName.All_Changing_Account_Failed,
                                                   "InfluencerContainer::read_stories-LoginRequired",
                                                   robot_api=RobotAPI.Instagrapi, target_account=infl_username,
                                                   bot_id=self.IA_engine.get_grapi_acc_id())
                    return False, False
        except Exception as e:
            EventRepository.register_event(EventType.Warning, EventCategory.Reading_Media,
                                           EventName.Reading_Story_Failed,
                                           "InfluencerContainer::read_stories-Others",
                                           robot_api=RobotAPI.Instagrapi, target_account=infl_username,
                                           bot_id=self.IA_engine.get_grapi_acc_id(), content=str(e))

            self.IA_engine.grapi_engine.logger.exception(e)
            sleep(random.randint(11, 24))
            relogin_flag = self.IA_engine.relogin_grapi.relogin("InfluencerContainer::read_stories-Others",
                                                                infl_username)
            if relogin_flag:
                try:
                    stories = self.IA_engine.grapi_engine.user_stories(infl_user_id)
                    EventRepository.register_event(EventType.Log, EventCategory.Reading_Media,
                                                   EventName.Reading_Story, "InfluencerContainer::read_stories-"
                                                                            "Others",
                                                   robot_api=RobotAPI.Instagrapi, target_account=infl_username,
                                                   bot_id=self.IA_engine.get_grapi_acc_id())
                except Exception as e:
                    EventRepository.register_event(EventType.Warning, EventCategory.Reading_Media,
                                                   EventName.Reading_Story_Failed,
                                                   "InfluencerContainer::read_stories-Others",
                                                   robot_api=RobotAPI.Instagrapi, target_account=infl_username,
                                                   bot_id=self.IA_engine.get_grapi_acc_id(), content=str(e))
                    return None, False
            else:
                login_flag = self.IA_engine.change_account()
                if login_flag:
                    try:
                        stories = self.IA_engine.grapi_engine.user_stories(infl_user_id)
                        EventRepository.register_event(EventType.Log, EventCategory.Reading_Media,
                                                       EventName.Reading_Story, "InfluencerContainer::read_stories-"
                                                                                "Others",
                                                       robot_api=RobotAPI.Instagrapi, target_account=infl_username,
                                                       bot_id=self.IA_engine.get_grapi_acc_id())
                    except Exception as e:
                        EventRepository.register_event(EventType.Warning, EventCategory.Reading_Media,
                                                       EventName.Reading_Story_Failed,
                                                       "InfluencerContainer::read_stories-Others",
                                                       robot_api=RobotAPI.Instagrapi, target_account=infl_username,
                                                       bot_id=self.IA_engine.get_grapi_acc_id(), content=str(e))
                        return None, False
                else:
                    EventRepository.register_event(EventType.Error, EventCategory.Authentication,
                                                   EventName.All_Changing_Account_Failed,
                                                   "InfluencerContainer::read_stories-Others",
                                                   robot_api=RobotAPI.Instagrapi, target_account=infl_username,
                                                   bot_id=self.IA_engine.get_grapi_acc_id())
                    return False, False

        if len(stories) > 0:
            reading_stories_flag = True
        for story in stories:
            if (datetime.now().astimezone(timezone.utc) - story.taken_at.astimezone(timezone.utc)) < \
                    timedelta(minutes=60) and story.taken_at.astimezone(timezone.utc) > \
                    self.infl_repo.get_start_monitoring(infl_id).astimezone(timezone.utc):
                if not self.infl_repo.is_exist_story_id(infl_id, story.id):
                    sleep(random.randint(3, 10))
                    # try:
                    #     self.robo_engine.story_seen([story.pk])
                    # except PleaseWaitFewMinutes as e:
                    #     warning_title = "Story_seen method error on " + infl_username + " stories"
                    #     warning_message = str(e)
                    #     write_warning_message(warning_title, warning_message)
                    #     sleep(random.randint(300, 360))  # Wait for 5-6 minutes
                    #     try:
                    #         self.robo_engine.logger.exception(e)
                    #         sleep(random.randint(30, 60))
                    #         self.robo_engine.relogin()
                    #         write_Log_message(
                    #             "IA robo has relogged in, and skips " + infl_username + "'story_seen method\n" +
                    #             infl_username + "'s stories quantity after relogging: " +
                    #             str(influencer.get_corresponding_stories_count()))
                    #         return
                    #         # self.robo_engine.update_client_settings(self.robo_engine.get_settings())
                    #     except Exception as e:
                    #         error_title = "ReLogin failed"
                    #         error_message = e
                    #         tb = e.__traceback__
                    #         error_line = tb.tb_lineno
                    #         write_error_message(error_title, error_message, error_line)
                    self.infl_repo.register_story(infl_id, story.id, story.taken_at.astimezone(timezone.utc))
                    mentioned_links = story.mentions
                    for link in mentioned_links:
                        mentioned_user = link.user
                        start_datetime = story.taken_at.astimezone(timezone.utc)
                        mentioned_username = mentioned_user.username
                        adv_acc_id_check = self.adv_repo.check_account_exists(mentioned_username)
                        adv_id_to_check = self.adv_repo.check_advertise_exists(infl_id,
                                                                               adv_acc_id_check)
                        if adv_id_to_check is None:  # story was new:
                            sleep(random.randint(2, 5))
                            self.create_advertise(mentioned_username, infl_id, start_datetime,
                                                  infl_username)
                        else:
                            self.adv_repo.plus_story_count(adv_id_to_check)

                sleep(random.randint(2, 6))
        return True, reading_stories_flag

    def create_advertise(self, mentioned_username, infl_id, start_datetime, influencer_username):
        try:
            mentioned_profile = self.IA_engine.loader_engine.check_profile_id(mentioned_username)
            mentioned_id = mentioned_profile.userid
            followers_count = mentioned_profile.followers
            mentioned_media_count = mentioned_profile.mediacount
        except Exception as e:
            EventRepository.register_event(EventType.Error, EventCategory.Reading_Media,
                                           EventName.Reading_Profile_Failed, "InfluencerContainer::create_advertise",
                                           robot_api=RobotAPI.Instaloader, target_account=mentioned_username,
                                           bot_id=self.IA_engine.get_loader_acc_id(), content=str(e))
            login_flag = self.IA_engine.change_account("Loader")
            if login_flag:
                try:
                    mentioned_profile = self.IA_engine.loader_engine.check_profile_id(mentioned_username)
                    mentioned_id = mentioned_profile.userid
                    followers_count = mentioned_profile.followers
                    mentioned_media_count = mentioned_profile.mediacount
                except Exception as e:
                    EventRepository.register_event(EventType.Error, EventCategory.Reading_Media,
                                                   EventName.Reading_Profile_Failed,
                                                   "InfluencerContainer::create_advertise",
                                                   robot_api=RobotAPI.Instaloader, target_account=mentioned_username,
                                                   bot_id=self.IA_engine.get_loader_acc_id(), content=str(e))
                    try:
                        mentioned_id = self.IA_engine.grapi_engine.user_id_from_username(mentioned_username)
                        mentioned_profile = self.IA_engine.grapi_engine.user_info(mentioned_id)
                        followers_count = mentioned_profile.follower_count
                        mentioned_media_count = mentioned_profile.media_count
                    except Exception as e:
                        EventRepository.register_event(EventType.Error, EventCategory.Reading_Media,
                                                       EventName.Reading_Profile_Failed,
                                                       "InfluencerContainer::create_advertise",
                                                       robot_api=RobotAPI.Instagrapi, target_account=mentioned_username,
                                                       bot_id=self.IA_engine.get_grapi_acc_id(), content=str(e))
                        return False
            else:
                EventRepository.register_event(EventType.Error, EventCategory.Authentication,
                                               EventName.All_Changing_Account_Failed,
                                               "InfluencerContainer::create_advertise",
                                               robot_api=RobotAPI.Instaloader, target_account=mentioned_username,
                                               bot_id=self.IA_engine.get_loader_acc_id())
                return False

        mentioned_full_name = mentioned_profile.full_name[:100]
        mentioned_biography = mentioned_profile.biography[:200]
        mention_privacy = Privacy.Private if mentioned_profile.is_private else Privacy.Public

        like_avrg, comment_avrg, engagement = self.calculate_statistical_properties(mentioned_id,
                                                                                    followers_count, mentioned_username)
        if like_avrg:
            new_mentioned_adv_account = Account(
                mentioned_username, mentioned_id, mentioned_full_name, mentioned_biography, followers_count,
                mentioned_media_count, like_avrg, comment_avrg, engagement, mention_privacy)
        else:
            new_mentioned_adv_account = Account(
                mentioned_username, mentioned_id, mentioned_full_name, mentioned_biography, followers_count,
                mentioned_media_count, 0, 0, 0, mention_privacy)

        account_item_id = self.adv_repo.check_account_exists(mentioned_username)
        if account_item_id:
            self.adv_repo.register_advertise(new_mentioned_adv_account, infl_id, start_datetime, followers_count,
                                             account_item_id)
        else:
            self.adv_repo.register_advertise(new_mentioned_adv_account, infl_id, start_datetime, followers_count)

    def calculate_statistical_properties(self, user_id, followers_count, username):
        SINCE = datetime.now().date()
        UNTIL = SINCE - timedelta(days=300)
        try:
            posts = self.IA_engine.grapi_engine.user_medias(user_id, 50)
        except Exception as e:
            EventRepository.register_event(EventType.Warning, EventCategory.Reading_Media,
                                           EventName.Reading_Post_Failed, "InfluencerContainer::calculate_statistical",
                                           robot_api=RobotAPI.Instagrapi, target_account=username,
                                           bot_id=self.IA_engine.get_grapi_acc_id(), content=str(e))

            self.IA_engine.grapi_engine.logger.exception(e)
            sleep(random.randint(10, 20))
            relogin_flag = self.IA_engine.relogin_grapi.relogin("InfluencerContainer::calculate_statistical",
                                                                username)
            if relogin_flag:
                try:
                    posts = self.IA_engine.grapi_engine.user_medias(user_id, 50)
                except Exception as e:
                    EventRepository.register_event(EventType.Warning, EventCategory.Reading_Media,
                                                   EventName.Reading_Post_Failed,
                                                   "InfluencerContainer::calculate_statistical",
                                                   robot_api=RobotAPI.Instagrapi, target_account=username,
                                                   bot_id=self.IA_engine.get_grapi_acc_id(), content=str(e))
            else:
                login_flag = self.IA_engine.change_account()
                if login_flag:
                    try:
                        posts = self.IA_engine.grapi_engine.user_medias(user_id, 50)
                    except Exception as e:
                        EventRepository.register_event(EventType.Warning, EventCategory.Reading_Media,
                                                       EventName.Reading_Post_Failed,
                                                       "InfluencerContainer::calculate_statistical",
                                                       robot_api=RobotAPI.Instagrapi, target_account=username,
                                                       bot_id=self.IA_engine.get_grapi_acc_id(), content=str(e))
                        return False, False, False
                else:
                    EventRepository.register_event(EventType.Warning, EventCategory.Reading_Media,
                                                   EventName.Reading_Post_Failed,
                                                   "InfluencerContainer::calculate_statistical",
                                                   robot_api=RobotAPI.Instagrapi, target_account=username,
                                                   bot_id=self.IA_engine.get_grapi_acc_id(), content=str(e))
                    return False, False, False
        likes_count = 0
        comments_count = 0
        post_count = 0
        # for post_item in takewhile(lambda p: p.taken_at.date() > UNTIL,
        #                            dropwhile(lambda p: p.taken_at.date() > SINCE, posts)):
        for post_item in posts:
            if datetime.now(timezone.utc) - post_item.taken_at.astimezone(timezone.utc) < timedelta(days=360):
                post_count += 1
                likes_count += post_item.like_count
                comments_count += post_item.comment_count
        if post_count == 0:
            return 0, 0, 0
        else:
            like_average = (likes_count * 1.0) / post_count
            comment_average = (comments_count * 1.0) / post_count
            activity_per_follower = ((like_average + comment_average) / followers_count) * 100
            return like_average, comment_average, activity_per_follower

    def update_statistical_properties(self):
        incomplete_infl_id = self.infl_repo.get_incomplete_infl()
        for infl_id in incomplete_infl_id:
            current_user_id = self.infl_repo.get_user_id(infl_id)
            followers_count = self.infl_repo.get_follower(infl_id)
            infl_username = self.infl_repo.get_username(infl_id)
            like_average, comment_average, engagement_rate = \
                self.calculate_statistical_properties(current_user_id, followers_count, infl_username)
            if like_average:
                self.infl_repo.update_statistical_properties(infl_id, like_average, comment_average, engagement_rate)

    def create_influencer(self, influencer_username, start_monitoring_datetime, finish_monitoring_datetime, price=None):
        try:
            inf_profile = self.IA_engine.loader_engine.check_profile_id(influencer_username)
            influencer_id = inf_profile.userid
            followers_count = inf_profile.followers
            media_count = inf_profile.mediacount
        except Exception as e:
            EventRepository.register_event(EventType.Error, EventCategory.Reading_Media,
                                           EventName.Reading_Profile_Failed, "InfluencerContainer::create_influencer",
                                           robot_api=RobotAPI.Instaloader, target_account=influencer_username,
                                           bot_id=self.IA_engine.get_loader_acc_id(), content=str(e))
            login_flag = self.IA_engine.change_account("Loader")
            if login_flag:
                try:
                    inf_profile = self.IA_engine.loader_engine.check_profile_id(influencer_username)
                    influencer_id = inf_profile.userid
                    followers_count = inf_profile.followers
                    media_count = inf_profile.mediacount
                except Exception as e:
                    EventRepository.register_event(EventType.Error, EventCategory.Reading_Media,
                                                   EventName.Reading_Profile_Failed,
                                                   "InfluencerContainer::create_influencer",
                                                   robot_api=RobotAPI.Instaloader,
                                                   bot_id=self.IA_engine.get_loader_acc_id(),
                                                   target_account=influencer_username,
                                                   content=str(e))
                    try:
                        influencer_id = self.IA_engine.grapi_engine.user_id_from_username(influencer_username)
                        inf_profile = self.IA_engine.grapi_engine.user_info(influencer_id)
                        followers_count = inf_profile.follower_count
                        media_count = inf_profile.media_count
                    except Exception as e:
                        EventRepository.register_event(EventType.Error, EventCategory.Reading_Media,
                                                       EventName.Reading_Profile_Failed,
                                                       "InfluencerContainer::create_influencer",
                                                       robot_api=RobotAPI.Instagrapi,
                                                       bot_id=self.IA_engine.get_grapi_acc_id(),
                                                       target_account=influencer_username,
                                                       content=str(e))
                        return False
            else:
                EventRepository.register_event(EventType.Error, EventCategory.Authentication,
                                               EventName.All_Changing_Account_Failed,
                                               "InfluencerContainer::read_stories-Others",
                                               robot_api=RobotAPI.Instaloader, target_account=influencer_username,
                                               bot_id=self.IA_engine.get_loader_acc_id())
                return False

        full_name = inf_profile.full_name[:100]
        biography = inf_profile.biography[:200]
        influencer_privacy = Privacy.Private if inf_profile.is_private else Privacy.Public

        like_average, comment_average, activity_per_follower = self.calculate_statistical_properties(influencer_id,
                                                                                                     followers_count,
                                                                                                     influencer_username)
        existed_account = self.infl_repo.username_account_exists(influencer_username)
        if existed_account is not None:
            self.infl_repo.Register_influencer(existed_account, start_monitoring_datetime, finish_monitoring_datetime,
                                               price)
        else:
            if not like_average:
                account = Account(influencer_username, influencer_id, full_name, biography, followers_count,
                                  media_count,
                                  0, 0, 0, influencer_privacy)
            else:
                account = Account(influencer_username, influencer_id, full_name, biography, followers_count,
                                  media_count,
                                  like_average, comment_average, activity_per_follower, influencer_privacy)

            self.infl_repo.Register_influencer(account, start_monitoring_datetime, finish_monitoring_datetime)

    def get_active_influencer_advertises(self, infl_id):
        return self.adv_repo.get_active_influencer_advertises(infl_id)

    def get_all_influencers_id(self):
        return self.infl_repo.get_all_influencers()

    def get_delta_time(self, adv_id):
        current_delta_time = datetime.now().astimezone(timezone.utc) - \
                             self.adv_repo.get_start_datetime(adv_id)
        mentioned_username = self.adv_repo.get_mentioned_username(adv_id)

        return current_delta_time, mentioned_username

    def create_xlsx_result(self, influencer_username):
        infl_id = self.infl_repo.check_influencer_exists(influencer_username)
        if not infl_id:
            return

        address = "OutputFiles/" + influencer_username + ".xlsx"
        workbook = xlsxwriter.Workbook(address)
        str_format = workbook.add_format({'font_name': 'B Homa', 'font_size': 14})
        str_format.set_center_across()
        str_format.set_align('center')
        int_format = workbook.add_format({'num_format': "Number"})
        int_format.set_center_across()

        influencer_sheet = workbook.add_worksheet("influencer_datas")
        influencer_sheet.right_to_left()
        influencer_sheet.write(0, 0, "شناسه اینفلوئنسر", str_format)
        influencer_sheet.write(0, 1, self.infl_repo.get_username(infl_id), str_format)
        influencer_sheet.write(1, 0, "نام اکانت", str_format)
        influencer_sheet.write(1, 1, self.infl_repo.get_full_name(infl_id), str_format)
        influencer_sheet.write(2, 0, "بیوگرافی", str_format)
        influencer_sheet.write(2, 1, self.infl_repo.get_bio(infl_id), str_format)
        influencer_sheet.write(3, 0, "تعداد فالور", str_format)
        influencer_sheet.write(3, 1, self.infl_repo.get_follower(infl_id), int_format)
        influencer_sheet.write(4, 0, "تعداد پست", str_format)
        influencer_sheet.write(4, 1, self.infl_repo.get_media_count(infl_id), int_format)
        influencer_sheet.write(5, 0, "میانگین لایک", str_format)
        influencer_sheet.write(5, 1, self.infl_repo.get_AVG_Like(infl_id), int_format)
        influencer_sheet.write(6, 0, "میانگین کامنت", str_format)
        influencer_sheet.write(6, 1, self.infl_repo.get_AVG_Comment(infl_id), int_format)
        influencer_sheet.write(7, 0, "EngagementRate", str_format)
        influencer_sheet.write(7, 1, self.infl_repo.get_engagement_rate(infl_id), int_format)

        mention_account_sheet = workbook.add_worksheet("mention_accounts")
        mention_account_sheet.right_to_left()
        date_format = workbook.add_format(
            {'num_format': '[$-fa-IR,16]yyyy/mm/dd', 'font_name': 'B Homa', 'font_size': 8})
        time_format = workbook.add_format({'num_format': 'h:mm'})

        mention_account_sheet.write(0, 0, "ردیف", str_format)
        mention_account_sheet.write(0, 1, "تاریخ تبلیغ", str_format)
        mention_account_sheet.write(0, 2, "ساعت تبلیغ", str_format)
        mention_account_sheet.write(0, 3, "شناسه پیج معرفی شده", str_format)
        mention_account_sheet.write(0, 4, "نام پیج معرفی شده", str_format)
        mention_account_sheet.write(0, 5, "بیوگرافی پیج معرفی شده", str_format)
        mention_account_sheet.write(0, 6, "تعداد پست", str_format)
        mention_account_sheet.write(0, 7, "تعداد استوری تبلیغ", str_format)
        mention_account_sheet.write(0, 8, "تعداد فالور قبل تبلیغ", str_format)
        mention_account_sheet.write(0, 9, "تعداد فالور بعد از 1 ساعت تبلیغ", str_format)
        mention_account_sheet.write(0, 10, "تعداد فالور بعد از 2 ساعت تبلیغ", str_format)
        mention_account_sheet.write(0, 11, "تعداد فالور بعد از 12 ساعت تبلیغ", str_format)
        mention_account_sheet.write(0, 12, "تعداد فالور بعد از 24 ساعت تبلیغ", str_format)
        mention_account_sheet.write(0, 13, "بازخورد", str_format)
        mention_account_sheet.write(0, 14, "وضعیت اکانت", str_format)
        mention_account_sheet.write(0, 15, "وضعیت تبلیغ", str_format)
        workbook.close()

        #### copy style
        workbook1 = openpyxl.load_workbook(address)
        inf_account_sheet = workbook1["influencer_datas"]
        men_account_sheet = workbook1["mention_accounts"]

        workbook_t = openpyxl.load_workbook("InputFiles/template.xlsx")
        inf_account_sheet_t = workbook_t["influencer_datas"]
        men_account_sheet_t = workbook_t["mention_accounts"]
        inf_account_sheet.column_dimensions["B"].width = 60.0
        inf_account_sheet.column_dimensions["A"].width = 40.0
        for i in range(1, 9):
            newcell = inf_account_sheet.cell(i, 1)
            cell = inf_account_sheet_t.cell(i, 1)
            newcell.value = cell.value
            self.format_copy(newcell, cell)
        for i in range(1, 9):
            newcell = inf_account_sheet.cell(i, 2)
            cell = inf_account_sheet_t.cell(i, 2)
            self.format_copy(newcell, cell)
        # mention account sheet
        men_account_sheet.column_dimensions["A"].width = 20.0
        men_account_sheet.column_dimensions["B"].width = 20.0
        men_account_sheet.column_dimensions["C"].width = 20.0
        men_account_sheet.column_dimensions["D"].width = 30.0
        men_account_sheet.column_dimensions["E"].width = 25.0
        men_account_sheet.column_dimensions["F"].width = 80.0
        men_account_sheet.column_dimensions["G"].width = 15.0
        men_account_sheet.column_dimensions["H"].width = 15.0
        men_account_sheet.column_dimensions["I"].width = 20.0
        men_account_sheet.column_dimensions["J"].width = 20.0
        men_account_sheet.column_dimensions["K"].width = 20.0
        men_account_sheet.column_dimensions["L"].width = 20.0
        men_account_sheet.column_dimensions["M"].width = 20.0
        men_account_sheet.column_dimensions["N"].width = 20.0
        men_account_sheet.column_dimensions["O"].width = 20.0
        men_account_sheet.column_dimensions["P"].width = 20.0
        men_account_sheet.column_dimensions["Q"].width = 20.0

        for j in range(1, 18):
            newcell1 = men_account_sheet.cell(1, j)
            cell1 = men_account_sheet_t.cell(1, j)
            newcell1.value = cell1.value
            self.format_copy(newcell1, cell1)

        workbook1.save(address)

    def format_copy(self, nc, c):
        nc.font = copy(c.font)
        nc.border = copy(c.border)
        nc.fill = copy(c.fill)
        nc.number_format = copy(c.number_format)
        nc.protection = copy(c.protection)
        nc.alignment = copy(c.alignment)

    def create_xlsx_report(self):
        for infl_id in self.get_all_influencers_id():
            file_name = self.infl_repo.get_username(infl_id) + ".xlsx"
            if not path.isdir("OutputFiles"):
                makedirs("OutputFiles")

            self.create_xlsx_result(self.infl_repo.get_username(infl_id))

            address = "OutputFiles/" + file_name
            workbook = openpyxl.load_workbook(address)
            mention_account_sheet = workbook["mention_accounts"]

            new_row = mention_account_sheet.max_row + 1
            columns_number = mention_account_sheet.max_column
            workbook_t = openpyxl.load_workbook("InputFiles/template.xlsx")

            template_file = openpyxl.load_workbook("InputFiles/template.xlsx")
            template_sheet = template_file["mention_accounts"]
            template_line = template_sheet.max_row

            if len(mention_account_sheet.tables) != 0:
                del mention_account_sheet.tables["Table1"]
            ##############################
            adv_ids = self.adv_repo.get_all_infl_advs(infl_id)
            for adv_id in adv_ids:
                if (new_row % 2) == 0:
                    template_line = 2
                else:
                    template_line = 3

                for i in range(1, columns_number + 1):
                    nc = mention_account_sheet.cell(new_row, i)
                    cc = template_sheet.cell(template_line, i)
                    self.format_copy(nc, cc)
                mention_account_name = self.adv_repo.get_mentioned_username(adv_id)
                story_date = self.adv_repo.get_start_datetime(adv_id).astimezone(pytz.timezone('Asia/Tehran')).date()
                changed_date = datetime.strptime(story_date.strftime("%Y-%m-%d"), "%Y-%m-%d")
                mention_account_sheet.cell(new_row, 2).value = changed_date
                mention_account_sheet.cell(new_row, 2).number_format = "[$-fa-IR,16]yyyy/mm/dd"
                mention_account_sheet.cell(new_row, 1).value = new_row - 1

                story_time = self.adv_repo.get_start_datetime(adv_id).astimezone(pytz.timezone('Asia/Tehran')).time()
                changed_time = datetime.strptime(story_time.strftime("%H-%M"), "%H-%M")
                mention_account_sheet.cell(new_row, 3).value = changed_time
                mention_account_sheet.cell(new_row, 3).number_format = 'h:mm'

                mention_account_sheet.cell(new_row, 4).value = mention_account_name
                mention_account_sheet.cell(new_row, 5).value = self.adv_repo.get_mentioned_username(adv_id)
                mention_account_sheet.cell(new_row, 6).value = self.adv_repo.get_mentioned_bio(adv_id)
                mention_account_sheet.cell(new_row, 7).value = self.adv_repo.get_mentioned_media_count(adv_id)
                mention_account_sheet.cell(new_row, 8).value = self.adv_repo.get_mentioned_story_count(adv_id)
                mention_account_sheet.cell(new_row, 9).value = self.adv_repo.get_follower_before(adv_id)
                mention_account_sheet.cell(new_row, 10).value = self.adv_repo.get_follower_after_1h(adv_id)
                mention_account_sheet.cell(new_row, 11).value = self.adv_repo.get_follower_after_2h(adv_id)
                mention_account_sheet.cell(new_row, 12).value = self.adv_repo.get_follower_after_12h(adv_id)
                mention_account_sheet.cell(new_row, 13).value = self.adv_repo.get_follower_after_24h(adv_id)
                mention_account_sheet.cell(new_row,
                                           14).value = self.adv_repo.get_follower_after_24h(adv_id) - \
                                                       self.adv_repo.get_follower_before(adv_id)

                if self.adv_repo.get_mentioned_privacy(adv_id) == Privacy.Public:
                    mention_account_sheet.cell(new_row, 15).value = "عمومی"  # public
                else:
                    mention_account_sheet.cell(new_row, 15).value = "خصوصی"

                if self.adv_repo.get_mentioned_status(adv_id) == Advertise_Status.Open:
                    mention_account_sheet.cell(new_row, 16).value = "باز"
                else:
                    mention_account_sheet.cell(new_row, 16).value = "بسته"
                new_row = new_row + 1

            endline = mention_account_sheet.max_row
            refs = "A1:P" + str(endline)

            tab = Table(displayName="Table1", ref=refs)
            mention_account_sheet.add_table(tab)
            workbook.save(address)

    def set_follower_after_1h(self, adv_id, followers_count):
        self.adv_repo.set_follower_after_1h(adv_id, followers_count)

    def set_follower_after_2h(self, adv_id, followers_count):
        self.adv_repo.set_follower_after_2h(adv_id, followers_count)

    def set_follower_after_12h(self, adv_id, followers_count):
        self.adv_repo.set_follower_after_12h(adv_id, followers_count)

    def set_follower_after_24h(self, adv_id, followers_count):
        self.adv_repo.set_follower_after_24h(adv_id, followers_count)
        performance = followers_count - self.adv_repo.get_follower_before(adv_id)
        self.adv_repo.set_performance(adv_id, performance)

    def get_follower_after_1h(self, adv_id):
        return self.adv_repo.get_follower_after_1h(adv_id)

    def get_follower_after_2h(self, adv_id):
        return self.adv_repo.get_follower_after_2h(adv_id)

    def get_follower_after_12h(self, adv_id):
        return self.adv_repo.get_follower_after_12h(adv_id)

    def get_follower_after_24h(self, adv_id):
        return self.adv_repo.get_follower_after_24h(adv_id)

    def set_advertise_status(self, adv_id, status):
        self.adv_repo.set_advertise_status(adv_id, status)
